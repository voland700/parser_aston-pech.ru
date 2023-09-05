from bs4 import BeautifulSoup
import urllib.request
import fake_useragent
import json
import requests
import pathlib
import os
import string
import random
import openpyxl
import datetime

class getAstonContent():
    ua = fake_useragent.UserAgent()
    def __init__(self):
        self.headers = {"User-Agent": self.ua.random}
        self.propertiesName = []
        self.products = {}

    def randomStr(self, nrandchars):
        alpha = string.ascii_letters + string.digits
        chars = ''.join(random.choice(alpha) for _ in range(nrandchars))
        return chars

    def get_files(self, link, folder='images'):
        response = requests.get(link, stream=True)
        ext = pathlib.Path(link).suffix
        if not os.path.isdir("upload"):
            os.mkdir("upload")
        if not os.path.isdir("upload/" + folder):
            os.mkdir("upload/" + folder)
        file_name = 'upload/' + folder + '/' + folder + '_' + self.randomStr(10) + ext
        file = open(file_name, 'bw')
        for chunk in response.iter_content(4096):
            file.write(chunk)
        return file_name


    def getLinksFromPageCategory(self, link):
        req = urllib.request.Request(link, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        if req.getcode() != 200:
            return False
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        links = []
        for a in soup.find('ul', class_='products').find_all('a', class_='btn'):
            links.append(a['href'])
        return links

    def getItem(self, url):
        item = dict()
        item['url'] = url
        item['title'] = None
        item['price'] = 0
        item['main'] = None
        item['more'] = ''
        item['description'] = ''
        item['video'] = ''
        item['json'] = None
        item['attributes'] = None

        req = urllib.request.Request(url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        if req.getcode() != 200:
            return False
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        itemContent = soup.find('div', class_='product')
        item['title'] = itemContent.find('span', {'class':'product_title'}).get_text(strip=True)
        item['price'] = itemContent.find('span', {'class':'woocommerce-Price-amount'}).get_text(strip=True)
        try:
            item['video'] = itemContent.find('div', {'class':'single_video'}).a['href']
        except:
            pass
        try:
            item['description'] = itemContent.find('div', {'class':'single_description'}).decode_contents().replace('\n', '').replace('\r', '')
        except:
            pass
        attributes = soup.find_all('tr', class_='woocommerce-product-attributes-item')
        if len(attributes) > 0:
            atrr = []
            for props in attributes:
                element= {
                    'label': props.find('th', class_='woocommerce-product-attributes-item__label').get_text(strip=True),
                    'value': props.find('td', class_='woocommerce-product-attributes-item__value').get_text(strip=True)
                }
                if element['label'] in self.propertiesName:
                    pass
                else:
                    self.propertiesName.append(element['label'])
                atrr.append(element)
            if len(atrr) > 0:
                forJson = {}
                key = 1
                for el in atrr:
                    forJson[key] = el
                    key += 1
                item['json'] = json.dumps(forJson, ensure_ascii=False)

            item['attributes'] = atrr
        try:
            mainImgLink = soup.find('div', class_='single_image').a['href']
            item['main'] = self.get_files(mainImgLink)
        except:
            pass
        moreImgLinks = soup.find('div', class_='single_thumbnails').find_all('a')
        if len(moreImgLinks) > 0:
            moreName = []
            for itemMore in moreImgLinks:
                try:
                    path = self.get_files(itemMore['href'], 'more')
                    moreName.append(path)
                except:
                    pass
            if len(moreName) > 0:
                item['more'] = ','.join(moreName)

        return item


    def getExel(self, data):
        book = openpyxl.Workbook()
        sheet = book.active

        sheet.cell(row=1, column=1).value = 'URL'
        sheet.cell(row=1, column=2).value = 'NAME'
        sheet.cell(row=1, column=3).value = 'PRICE'
        sheet.cell(row=1, column=4).value = 'DESCRIPTION'
        sheet.cell(row=1, column=5).value = 'MAIN_IMG'
        sheet.cell(row=1, column=6).value = 'MORE_IMG'
        sheet.cell(row=1, column=7).value = 'VIDEO'
        sheet.cell(row=1, column=8).value = 'JSON_PROPERTIES'

        maxColumn = 9 + len(self.propertiesName)
        if maxColumn > 0:
            propNumber = 9
            for name in self.propertiesName:
                sheet.cell(row=1, column=propNumber).value = name
                propNumber += 1

        rowNumber = 2

        for item in data:
            sheet.cell(row=rowNumber, column=1).value = item['url']
            sheet.cell(row=rowNumber, column=2).value = item['title']
            sheet.cell(row=rowNumber, column=3).value = item['price']
            sheet.cell(row=rowNumber, column=4).value = item['description']
            sheet.cell(row=rowNumber, column=5).value = item['main']
            sheet.cell(row=rowNumber, column=6).value = item['more']
            sheet.cell(row=rowNumber, column=7).value = item['video']
            sheet.cell(row=rowNumber, column=8).value = item['json']
            if len(item['attributes']) > 0:
                for col in range(9, maxColumn):
                    for atr in item['attributes']:
                        if sheet.cell(row=1, column=col).value == atr['label']:
                            sheet.cell(row=rowNumber, column=col).value = atr['value']
                        else:
                            pass
            rowNumber+= 1

        uniq_filename = str(datetime.datetime.now().date()) + '_' + str(datetime.datetime.now().time()).replace(':','_').replace('.','_')
        book.save('resault_'+uniq_filename+'.xlsx')
        book.close()



    def getVideo(self,  url):
        req = urllib.request.Request(url, data=None, headers=self.headers)
        req = urllib.request.urlopen(req)
        html = req.read()
        soup = BeautifulSoup(html, 'html.parser')
        itemContent = soup.find('div', class_='product')
        video = itemContent.find('div', {'class': 'single_video'}).a['href']

        return video


print('*********************************')